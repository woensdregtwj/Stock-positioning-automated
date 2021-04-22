"""Creates the output file"""

from stocks_dataframe import StockData

import datetime as dt
import openpyxl as pyxl
from openpyxl.styles import Font as Font


class FileOutputter:
    def __init__(self, stocks_data):
        self.data = [
            stock for stocks_list in stocks_data.values() for stock in stocks_list
        ]

    def output_data(self):
        wb = pyxl.Workbook()
        ws = wb.active

        self.__create_headers(ws)

        current_row = 2
        for stock in self.data:
            for key, value in stock.__dict__.items():
                for col in range(1, ws.max_column + 1):
                    header = ws.cell(row=1, column=col).value
                    if header == key.title():
                        cell_value = ws.cell(row=current_row, column=col)
                        cell_value.value = value
                        continue
            current_row += 1
                    # TODO - elif it is like profit, then create formula instead of data!!!

        extract_date = dt.datetime.now().strftime('%Y%m%d %H%M')
        wb.save(f"Stock Calculations - {extract_date}.xlsx")

    def __create_headers(self, ws):
        header_font = Font(size=14, bold=True)
        column = 1
        for header in StockData.__annotations__.keys():
            current_cell = ws.cell(row=1, column=column)
            current_cell.value = header.title()
            current_cell.font = header_font

            column += 1






